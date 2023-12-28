import csv
import json
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def do_check_file_exists(path_file):
    if os.path.exists(path_file):
        print(f"The file '{path_file}' exists.")
    else:
        print(f"The file '{path_file}' does not exist.")
        exit()


def convert_csv_to_xlsx(file_json, file_csv):
    try:
        with open(file_json, 'r') as file:
            data_json = json.load(file)

        # Read CSV file using pandas
        data = pd.read_csv(file_csv)

        # Create a new Workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = data_json["metadata"]["sheet-name"]  # Set sheet name

        # Set font style (Alignment, PatternFill and Font)
        align_data = Alignment(
            horizontal=data_json["header"]["cell-align-h"],
            vertical=data_json["header"]["cell-align-v"]
        )
        align_header = Alignment(
            horizontal=data_json["header"]["cell-align-h"],
            vertical=data_json["header"]["cell-align-v"]
        )
        fill_data = PatternFill(
            end_color=data_json["data"]["background-color"],
            fill_type=data_json["data"]["background-fill-type"],
            start_color=data_json["data"]["background-color"]
        )
        fill_header = PatternFill(
            end_color=data_json["header"]["background-color"],
            fill_type=data_json["header"]["background-fill-type"],
            start_color=data_json["header"]["background-color"]
        )
        font_data = Font(
            bold=data_json["data"]["font-bold"],
            color=data_json["data"]["font-color"],
            italic=data_json["data"]["font-italic"],
            name=data_json["data"]["font-name"],
            size=data_json["data"]["font-size"]
        )
        font_header = Font(
            bold=data_json["header"]["font-bold"],
            color=data_json["header"]["font-color"],
            italic=data_json["header"]["font-italic"],
            name=data_json["header"]["font-name"],
            size=data_json["header"]["font-size"]
        )

        # Apply header configurations
        if data_json["metadata"]["has-headers"]:
            for col_idx, col in enumerate(data.columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col)
                cell.alignment = align_header
                cell.fill = fill_header
                cell.font = font_header

        # Apply data configurations
        with open(file_csv) as f:
            reader = csv.reader(f, delimiter=',')
            for row_index, row in enumerate(reader, start=1):
                for column_index, cell_value in enumerate(row, start=1):
                    if row_index != (1 if data_json["metadata"]["has-headers"] else 0):
                        cell = ws.cell(row=row_index, column=column_index, value=cell_value)
                        cell.alignment = align_data
                        cell.fill = fill_data
                        cell.font = font_data

        # Save the workbook as XLSX file
        wb.save(file_csv.split('.')[0] + ".xlsx")

        print(f"CSV file '{file_csv}' successfully converted to XLSX file '{file_csv.split('.')[0] + ".xlsx"}'")

    except Exception as e:
        print("An error occurred:", e)
